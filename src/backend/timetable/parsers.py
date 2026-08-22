"""把教务系统导出的 Excel 解析成结构化的课程 / 上课时段。

支持三种导出格式，上传时自动识别：

1. ``undergraduate`` 本科选课清单
   一行一门课，「上课时间」形如 ``星期一第1-2节{1-16周};星期四第3-4节{1-15周(单)}``，
   「教学地点」按 ``;`` 与时间段一一对应。

2. ``graduate`` 研究生课表清单
   一行一门课，「上课时间」形如 ``3-18周： 周一 5-6节  ,2周： 周四 3-4节``，
   周次写在前面，后面可以跟多个「星期 + 节次」，「上课教室」按 ``,`` 对应。

3. ``grid`` 网格课表
   行是节次、列是星期，单元格里一行一门课：
   ``中国古代文学（二）/(1-2节)1-16周/ J-506/林雪云/2024级……1班/56``，
   同一格里换行分隔的多条是并列的可选课程（比如体育选项课）。

解析全程只在内存里进行，不写磁盘、不入库。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

from .constants import CN_NUMBERS, MAX_SECTION, MAX_WEEK, WEEKDAY_MAP


class ParseError(Exception):
    """文件读不了或者认不出格式时抛出，视图层翻译成 400。"""


# --------------------------------------------------------------------------- #
# 结果结构
# --------------------------------------------------------------------------- #


@dataclass
class Session:
    """一个上课时段：周几、第几节到第几节、哪些周、在哪上。"""

    weekday: int
    start_section: int
    end_section: int
    weeks: list[int]
    weeks_text: str = ""
    location: str = ""

    def as_dict(self, key: str) -> dict:
        return {
            "key": key,
            "weekday": self.weekday,
            "start_section": self.start_section,
            "end_section": self.end_section,
            "weeks": self.weeks,
            "weeks_text": self.weeks_text,
            "location": self.location,
        }


@dataclass
class Course:
    name: str
    teacher: str = ""
    course_id: str = ""
    class_name: str = ""
    category: str = ""
    credit: str = ""
    sessions: list[Session] = field(default_factory=list)

    @property
    def group_key(self) -> tuple:
        return (self.name, self.teacher, self.course_id)

    def as_dict(self, index: int) -> dict:
        return {
            "key": f"c{index}",
            "name": self.name,
            "teacher": self.teacher,
            "course_id": self.course_id,
            "class_name": self.class_name,
            "category": self.category,
            "credit": self.credit,
            "sessions": [
                session.as_dict(f"c{index}s{i}") for i, session in enumerate(self.sessions)
            ],
        }


@dataclass
class ParseResult:
    fmt: str
    fmt_label: str
    courses: list[Course]
    warnings: list[str] = field(default_factory=list)
    title: str = ""


FORMAT_LABELS = {
    "undergraduate": "本科选课清单",
    "graduate": "研究生课表清单",
    "grid": "网格课表",
}


# --------------------------------------------------------------------------- #
# 读文件
# --------------------------------------------------------------------------- #


def read_rows(stream, filename: str) -> list[list[str]]:
    """读第一张工作表，返回全是字符串的二维表。"""
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if suffix in {"xlsx", "xlsm"}:
        return _read_xlsx(stream)
    if suffix == "xls":
        return _read_xls(stream)
    raise ParseError("只支持 .xlsx / .xlsm / .xls 文件")


def _read_xlsx(stream) -> list[list[str]]:
    import warnings

    import openpyxl

    try:
        with warnings.catch_warnings():
            # 教务系统导出的文件常常没有默认样式表，openpyxl 每次都会 warn 一句
            warnings.simplefilter("ignore")
            workbook = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl 的异常类型很杂，统一翻译
        raise ParseError("这个 .xlsx 打不开，可能已损坏或不是真正的 Excel 文件") from exc

    try:
        sheet = workbook[workbook.sheetnames[0]]
        return [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_xls(stream) -> list[list[str]]:
    import xlrd

    try:
        book = xlrd.open_workbook(file_contents=stream.read())
    except Exception as exc:
        raise ParseError(
            "这个 .xls 打不开。部分教务系统导出的其实是网页文件，"
            "请先用 Excel 另存为 .xlsx 再上传"
        ) from exc

    sheet = book.sheet_by_index(0)
    return [
        [_cell_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        for r in range(sheet.nrows)
    ]


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # xlrd 把「56」读成 56.0，直接 str() 会变成 "56.0"
        value = int(value)
    # \xa0 是导出文件里常见的不换行空格，会让「数据库管理系统 」这类名字带一个尾巴
    return str(value).replace("\xa0", " ").strip()


# --------------------------------------------------------------------------- #
# 周次 / 时间文本
# --------------------------------------------------------------------------- #

_DASH = r"[-－—~～]"
_PARITY_RE = re.compile(r"[（(]\s*(单|双)\s*[）)]")


def parse_weeks(text: str, parity: str | None = None) -> list[int]:
    """``1-16周`` / ``1-15周(单)`` / ``3,5,7,9周`` / ``1-2,5-8周`` -> 周次列表。"""
    text = (text or "").replace("周", "").strip()
    if not text:
        return []

    weeks: set[int] = set()

    for item in re.split(r"[,，、]", text):
        item = item.strip()
        if not item:
            continue

        item_parity = parity
        mark = _PARITY_RE.search(item)
        if mark:
            item_parity = mark.group(1)
            item = (item[: mark.start()] + item[mark.end() :]).strip()

        parts = [part for part in re.split(rf"\s*{_DASH}\s*", item) if part]
        try:
            if len(parts) >= 2:
                begin, end = int(parts[0]), int(parts[1])
            else:
                begin = end = int(parts[0])
        except ValueError:
            continue

        if begin > end:
            begin, end = end, begin

        for week in range(max(begin, 1), min(end, MAX_WEEK) + 1):
            if item_parity == "单" and week % 2 == 0:
                continue
            if item_parity == "双" and week % 2 == 1:
                continue
            weeks.add(week)

    return sorted(weeks)


def _clamp_sections(start: str | int, end: str | int | None) -> tuple[int, int] | None:
    try:
        begin = int(start)
        finish = int(end) if end not in (None, "") else begin
    except (TypeError, ValueError):
        return None
    if begin > finish:
        begin, finish = finish, begin
    if begin < 1 or finish > MAX_SECTION:
        return None
    return begin, finish


# ---- 本科：星期一第1-2节{1-16周} ---- #

_UNDERGRAD_SLOT_RE = re.compile(
    rf"星期([一二三四五六日天])\s*第\s*(\d+)\s*(?:{_DASH}\s*(\d+))?\s*节\s*[{{｛]([^}}｝]*)[}}｝]"
)

# ---- 研究生：3-18周： 周一 5-6节 ---- #

_GRAD_HEADER_RE = re.compile(
    rf"((?:\d+(?:\s*{_DASH}\s*\d+)?(?:\s*[（(](?:单|双)[）)])?)"
    rf"(?:\s*[,，、]\s*\d+(?:\s*{_DASH}\s*\d+)?(?:\s*[（(](?:单|双)[）)])?)*)"
    r"\s*周\s*(?:[（(](单|双)[）)])?\s*[：:]"
)
_GRAD_SLOT_RE = re.compile(
    rf"周([一二三四五六日天])\s*(\d+)\s*(?:{_DASH}\s*(\d+))?\s*节"
)


def parse_time_text(text: str) -> tuple[list[Session], list[str]]:
    """解析「上课时间」列，自动区分本科 / 研究生两种写法。"""
    text = (text or "").strip()
    if not text or text.lower() == "nan":
        return [], []

    if _UNDERGRAD_SLOT_RE.search(text):
        return _parse_undergraduate_time(text)
    if _GRAD_SLOT_RE.search(text):
        return _parse_graduate_time(text)
    return [], [f"无法识别的上课时间：{text}"]


def _parse_undergraduate_time(text: str) -> tuple[list[Session], list[str]]:
    sessions, warnings = [], []

    for match in _UNDERGRAD_SLOT_RE.finditer(text):
        sections = _clamp_sections(match.group(2), match.group(3))
        if sections is None:
            warnings.append(f"节次超出范围，已跳过：{match.group(0)}")
            continue

        weeks = parse_weeks(match.group(4))
        if not weeks:
            warnings.append(f"没解析出周次，已跳过：{match.group(0)}")
            continue

        sessions.append(
            Session(
                weekday=WEEKDAY_MAP[match.group(1)],
                start_section=sections[0],
                end_section=sections[1],
                weeks=weeks,
                weeks_text=match.group(4).strip(),
            )
        )

    return sessions, warnings


def _parse_graduate_time(text: str) -> tuple[list[Session], list[str]]:
    """研究生格式的周次写在前面，一个周次头可以带多个「星期 + 节次」。"""
    headers = [
        (match.end(), parse_weeks(match.group(1), match.group(2)), match.group(0).rstrip("：: "))
        for match in _GRAD_HEADER_RE.finditer(text)
    ]

    sessions, warnings = [], []

    for match in _GRAD_SLOT_RE.finditer(text):
        weeks: list[int] = []
        weeks_text = ""
        # 就近归属：取该时段之前最后一个周次头
        for end, header_weeks, header_text in headers:
            if end <= match.start():
                weeks, weeks_text = header_weeks, header_text
            else:
                break

        if not weeks:
            warnings.append(f"找不到对应的周次，已跳过：{match.group(0).strip()}")
            continue

        sections = _clamp_sections(match.group(2), match.group(3))
        if sections is None:
            warnings.append(f"节次超出范围，已跳过：{match.group(0).strip()}")
            continue

        sessions.append(
            Session(
                weekday=WEEKDAY_MAP[match.group(1)],
                start_section=sections[0],
                end_section=sections[1],
                weeks=weeks,
                weeks_text=weeks_text,
            )
        )

    return sessions, warnings


def split_locations(text: str) -> list[str]:
    """教学地点：本科用 ``;``、研究生用 ``,`` 分隔，这里一起处理。"""
    if not text or text.lower() == "nan":
        return []
    return [part.strip() for part in re.split(r"[;；,，]", text) if part.strip()]


def pick_location(locations: list[str], index: int) -> str:
    """地点数量和时间段数量不总是一一对应，这里做兜底。"""
    if not locations:
        return ""
    if index < len(locations):
        return locations[index]
    if len(set(locations)) == 1:
        return locations[0]
    return locations[-1]


# --------------------------------------------------------------------------- #
# 清单格式（本科 / 研究生）
# --------------------------------------------------------------------------- #

# 每个字段可以有多个候选列名，按顺序取第一个命中的
LIST_COLUMNS = {
    "name": ["课程名称", "课程"],
    "teacher": ["教师姓名", "任课教师", "授课教师"],
    "time": ["上课时间"],
    "location": ["教学地点", "上课教室", "上课地点"],
    "course_id": ["课程编号", "课程代码"],
    "class_name": ["教学班名称", "教学班"],
    "category": ["课程类别", "课程性质"],
    "credit": ["学分"],
}


def _find_header_row(rows: list[list[str]]) -> int | None:
    for index, row in enumerate(rows[:10]):
        if any(cell == "上课时间" for cell in row):
            return index
    return None


def _column_map(header: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for field_name, candidates in LIST_COLUMNS.items():
        for candidate in candidates:
            if candidate in header:
                mapping[field_name] = header.index(candidate)
                break
    return mapping


def _parse_list(rows: list[list[str]], header_index: int) -> ParseResult:
    header = rows[header_index]
    columns = _column_map(header)

    if "name" not in columns or "time" not in columns:
        raise ParseError("表头里找不到「课程名称」或「上课时间」列，请确认导出的是完整表格")

    fmt = "graduate" if "上课教室" in header else "undergraduate"

    def cell(row: list[str], key: str) -> str:
        index = columns.get(key)
        if index is None or index >= len(row):
            return ""
        return row[index]

    courses: list[Course] = []
    warnings: list[str] = []

    for row in rows[header_index + 1 :]:
        name = cell(row, "name")
        time_text = cell(row, "time")
        if not name or not time_text:
            continue

        teacher, teacher_no = _split_teacher(cell(row, "teacher"))

        sessions, row_warnings = parse_time_text(time_text)
        warnings.extend(f"{name}：{message}" for message in row_warnings)
        if not sessions:
            continue

        locations = split_locations(cell(row, "location"))
        for index, session in enumerate(sessions):
            session.location = pick_location(locations, index)

        courses.append(
            Course(
                name=name,
                teacher=teacher + (f"（{teacher_no}）" if teacher_no else ""),
                course_id=cell(row, "course_id"),
                class_name=cell(row, "class_name"),
                category=cell(row, "category"),
                credit=cell(row, "credit"),
                sessions=sessions,
            )
        )

    return ParseResult(fmt=fmt, fmt_label=FORMAT_LABELS[fmt], courses=_merge(courses),
                       warnings=warnings)


_TEACHER_NO_RE = re.compile(r"^(.*?)(\d{3,})$")


def _split_teacher(text: str) -> tuple[str, str]:
    """研究生表里的任课教师带工号，例如「谢钦2075」。"""
    text = (text or "").strip()
    match = _TEACHER_NO_RE.match(text)
    if match and match.group(1):
        return match.group(1).strip(), match.group(2)
    return text, ""


# --------------------------------------------------------------------------- #
# 网格课表
# --------------------------------------------------------------------------- #

_GRID_WEEKDAY_RE = re.compile(r"^(?:星期|周)([一二三四五六日天])$")

# 中国古代文学（二）/(1-2节)1-16周/ J-506/林雪云/2024级……/56
_GRID_ENTRY_RE = re.compile(
    rf"^(?P<name>[^/]+?)\s*/\s*[（(]\s*(?P<start>\d+)\s*(?:{_DASH}\s*(?P<end>\d+))?\s*节\s*[）)]"
    r"\s*(?P<weeks>[^/]*)/\s*(?P<location>[^/]*?)\s*(?:/\s*(?P<teacher>[^/]*?)\s*)?(?:/.*)?$"
)

_GRID_SECTION_LABEL_RE = re.compile(r"^第(.+?)节$")


def _grid_weekday_columns(rows: list[list[str]]) -> tuple[int, dict[int, int]] | None:
    for index, row in enumerate(rows[:10]):
        columns = {}
        for column, cell in enumerate(row):
            match = _GRID_WEEKDAY_RE.match(cell.strip())
            if match:
                columns[column] = WEEKDAY_MAP[match.group(1)]
        if len(columns) >= 3:
            return index, columns
    return None


def _sections_from_label(label: str) -> tuple[int, int] | None:
    """「第一二节」「第十一十二节」-> (1, 2) / (11, 12)。"""
    match = _GRID_SECTION_LABEL_RE.match(label.strip())
    if not match:
        return None

    body = match.group(1)
    if body in CN_NUMBERS:  # 「第五节」
        return CN_NUMBERS[body], CN_NUMBERS[body]
    # 「第一二节」「第十一十二节」：切成两半，从左往右第一个两边都认识的切法就是答案
    for split in range(1, len(body)):
        left, right = body[:split], body[split:]
        if left in CN_NUMBERS and right in CN_NUMBERS:
            return CN_NUMBERS[left], CN_NUMBERS[right]
    return None


def _parse_grid(rows: list[list[str]], header_index: int, columns: dict[int, int]) -> ParseResult:
    courses: list[Course] = []
    warnings: list[str] = []

    for row in rows[header_index + 1 :]:
        # 行首两列通常是「上午 / 第一二节」，作为节次的兜底
        label_sections = None
        for cell in row[: min(2, len(row))]:
            label_sections = label_sections or _sections_from_label(cell)

        for column, weekday in columns.items():
            if column >= len(row):
                continue
            for line in re.split(r"[\r\n]+", row[column]):
                line = line.strip()
                if not line:
                    continue

                course = _parse_grid_entry(line, weekday, label_sections)
                if course is None:
                    warnings.append(f"这一格看不懂，已跳过：{line}")
                    continue
                courses.append(course)

    if not courses:
        raise ParseError("认出了网格课表，但没解析出任何课程，请确认单元格里有课程信息")

    return ParseResult(fmt="grid", fmt_label=FORMAT_LABELS["grid"], courses=_merge(courses),
                       warnings=warnings)


def _parse_grid_entry(line: str, weekday: int, label_sections: tuple[int, int] | None):
    match = _GRID_ENTRY_RE.match(line)

    if match:
        sections = _clamp_sections(match.group("start"), match.group("end"))
        weeks_text = (match.group("weeks") or "").strip()
        name = match.group("name").strip()
        location = (match.group("location") or "").strip()
        teacher = (match.group("teacher") or "").strip()
    else:
        # 宽松兜底：至少要能找到周次，节次退回行标题
        weeks_match = re.search(rf"(\d+(?:\s*{_DASH}\s*\d+)?[^/]*?)周", line)
        if not weeks_match:
            return None
        sections = label_sections
        weeks_text = weeks_match.group(0).strip()
        parts = [part.strip() for part in line.split("/")]
        name = parts[0]
        location = parts[2] if len(parts) > 2 else ""
        teacher = parts[3] if len(parts) > 3 else ""

    if sections is None:
        return None

    weeks = parse_weeks(weeks_text)
    if not weeks:
        return None

    return Course(
        name=name,
        teacher=teacher,
        sessions=[
            Session(
                weekday=weekday,
                start_section=sections[0],
                end_section=sections[1],
                weeks=weeks,
                weeks_text=weeks_text,
                location=location,
            )
        ],
    )


# --------------------------------------------------------------------------- #
# 合并 + 入口
# --------------------------------------------------------------------------- #


def _merge(courses: list[Course]) -> list[Course]:
    """同名同老师的课合并成一门，时段按周几 / 节次排序。"""
    merged: dict[tuple, Course] = {}

    for course in courses:
        existing = merged.get(course.group_key)
        if existing is None:
            merged[course.group_key] = course
            continue
        for session in course.sessions:
            if not any(_same_session(session, other) for other in existing.sessions):
                existing.sessions.append(session)

    result = list(merged.values())
    for course in result:
        course.sessions.sort(key=lambda s: (s.weekday, s.start_section, s.weeks[:1]))
    result.sort(key=lambda c: (c.sessions[0].weekday, c.sessions[0].start_section, c.name))
    return result


def _same_session(left: Session, right: Session) -> bool:
    return (
        left.weekday == right.weekday
        and left.start_section == right.start_section
        and left.end_section == right.end_section
        and left.weeks == right.weeks
        and left.location == right.location
    )


def parse_workbook(stream, filename: str) -> ParseResult:
    """上传文件 -> 解析结果。认不出格式会抛 ParseError。"""
    rows = read_rows(stream, filename)
    if not rows:
        raise ParseError("文件里没有任何内容")

    header_index = _find_header_row(rows)
    if header_index is not None:
        result = _parse_list(rows, header_index)
    else:
        grid = _grid_weekday_columns(rows)
        if grid is None:
            raise ParseError(
                "认不出这个课表格式。目前支持教务系统的「选课清单」和「课表」两种导出，"
                "请确认没有手工删掉表头"
            )
        result = _parse_grid(rows, grid[0], grid[1])
        result.title = _grid_title(rows, grid[0])

    if not result.courses:
        raise ParseError("没有解析出任何课程，请确认文件里有排课数据")

    return result


def _grid_title(rows: list[list[str]], header_index: int) -> str:
    """网格课表第一行通常写着「2025-2026年第1学期」和班级名。"""
    for row in rows[:header_index]:
        parts = []
        for cell in row:
            cell = cell.strip()
            if cell and cell not in parts:
                parts.append(cell)
        if parts:
            return " ".join(parts)[:120]
    return ""
